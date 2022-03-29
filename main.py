import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.utils.cell import get_column_letter

from PIL import Image
import numpy as np


def tohex(array):
    array = np.asarray(array, dtype='uint32')
    return ((array[:, :, 0] << 16) + (array[:, :, 1] << 8) + array[:, :, 2])


def cord(row, column):
    return get_column_letter(column)+str(row)


def rgb_to_hex(rgb):
    return "{:02x}{:02x}{:02x}".format(arr[0].tostring()[0], arr[0].tostring()[1], arr[0].tostring()[2])


# worksheet.column_dimensions[get_column_letter(i)].width = defaultColWidth
wb = openpyxl.Workbook()
ws = wb.active

image = Image.open("./image.png")
arr = np.array(image)
counter = 0
for r, i in enumerate(arr):
    for c, a in enumerate(i):

        ws[cord(r+1, c+1)].fill = PatternFill(
            start_color=Color(rgb_to_hex(a)), fill_type='solid')
        counter += 1
        print(f"Done: {counter/(arr.size/3)*100}%")
wb.save('fill.xlsx')
